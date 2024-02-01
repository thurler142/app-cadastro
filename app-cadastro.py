from PySimpleGUI import PySimpleGUI as sg
import openpyxl
from openpyxl.styles import Font, Alignment
sg.theme('DarkAmber')
layout = [
    [sg.Text('Usuario:')],
    [sg.InputText(key='user', clear = True)],
    [sg.Text('E-mail:')],
    [sg.InputText(key='email')],
    [sg.Text('Senha:')],
    [sg.InputText(key='password', password_char='*')],
    [sg.Text('Confirmação de senha:')],
    [sg.InputText(key='password_check', password_char='*')],
    [sg.Button('cadastrar', key='cadastro')],
    [sg.Text('',key='mensagem')]
]
window = sg.Window('Sistema', layout)
workbook = openpyxl.Workbook()
workbook.create_sheet('cadastros')
sheet_cadastro = workbook['cadastros']
sheet_cadastro['A1'].value = 'Usuários'
sheet_cadastro['B1'].value = 'Senhas'
sheet_cadastro['C1'].value = 'Emails'

def Cadastro(workbook, usuario, senha, email):
    
    sheet_cadastro = workbook['cadastros'] if 'cadastros' in workbook.sheetnames else workbook.create_sheet('usuarios_cadastrados')
    sheet_cadastro.append([usuario, senha, email])
    
while True:
    event, values = window.read()
    if event == sg.WIN_CLOSED:
        break
    elif event == 'cadastro':
        senha_correta = values['password_check']
        usuario = values['user']
        senha = values['password']
        email = values['email']
        if senha == senha_correta:
            window['mensagem'].update(f'Usuário {usuario} cadastrado com sucesso!')
            Cadastro(workbook, usuario, senha, email)
            
        else:
            window['mensagem'].update('As senhas devem ser iguais')                        
workbook.save('cadastrados.xlsx')

